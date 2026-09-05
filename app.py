
import os
from datetime import datetime, timedelta
from io import BytesIO
from zoneinfo import ZoneInfo

from flask import Flask, request, jsonify, send_file, Response
import hashlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from sqlalchemy import create_engine, Column, Integer, String, Float, DateTime, Boolean, select
from sqlalchemy.orm import declarative_base, sessionmaker

APP_TZ = ZoneInfo("America/Lima")
API_KEY = os.environ.get("ESP32_API_KEY", "CAMBIA_ESTA_CLAVE")

database_url = os.environ.get("DATABASE_URL", "sqlite:///comedero_cloud.db")

# Neon entrega normalmente postgresql://...
if database_url.startswith("postgresql://"):
    database_url = database_url.replace("postgresql://", "postgresql+psycopg://", 1)
elif database_url.startswith("postgres://"):
    database_url = database_url.replace("postgres://", "postgresql+psycopg://", 1)

connect_args = {}
if database_url.startswith("sqlite"):
    connect_args = {"check_same_thread": False}

engine = create_engine(
    database_url,
    pool_pre_ping=True,
    connect_args=connect_args
)

SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

class Medicion(Base):
    __tablename__ = "mediciones"

    id = Column(Integer, primary_key=True, autoincrement=True)
    mascota_id = Column(String(60), nullable=False, index=True)
    nombre = Column(String(100), nullable=False)
    fecha_hora = Column(DateTime, nullable=False, index=True)
    peso = Column(Float, nullable=False)

class ReinicioSesion(Base):
    __tablename__ = "reinicios_sesion"

    id = Column(Integer, primary_key=True, autoincrement=True)
    mascota_id = Column(String(60), nullable=False, unique=True, index=True)
    fecha_hora = Column(DateTime, nullable=False)

# Base.metadata.create_all movido después de los modelos

class Mascota(Base):
    __tablename__ = "mascotas"

    id = Column(Integer, primary_key=True, autoincrement=True)
    mascota_id = Column(String(60), unique=True, nullable=False, index=True)
    nombre = Column(String(100), nullable=False)
    clave_hash = Column(String(255), nullable=False)
    visible = Column(Boolean, default=True)
    fecha_creacion = Column(DateTime, nullable=False)


class Administrador(Base):
    __tablename__ = "administrador"

    id = Column(Integer, primary_key=True, autoincrement=True)
    clave_maestra_hash = Column(String(255), nullable=False)


class Configuracion(Base):
    __tablename__ = "configuracion"

    id = Column(Integer, primary_key=True, autoincrement=True)
    solicitar_clave_usuario = Column(Boolean, default=True)


def hash_clave(clave):
    return hashlib.sha256(str(clave).encode()).hexdigest()


Base.metadata.create_all(engine)


def inicializar_plataforma():
    with SessionLocal() as db:
        pet = db.execute(
            select(Mascota).where(Mascota.mascota_id == "gato01")
        ).scalar_one_or_none()

        if pet is None:
            pet = db.execute(
                select(Mascota).where(Mascota.nombre == "Bimbolete")
            ).scalar_one_or_none()

        if pet is None:
            db.add(Mascota(
                mascota_id="gato01",
                nombre="Bimbolete",
                clave_hash=hash_clave("1234"),
                visible=True,
                fecha_creacion=now_lima()
            ))

        admin = db.execute(select(Administrador)).scalar_one_or_none()
        if admin is None:
            db.add(Administrador(
                clave_maestra_hash=hash_clave("9999")
            ))

        cfg = db.execute(select(Configuracion)).scalar_one_or_none()
        if cfg is None:
            db.add(Configuracion(solicitar_clave_usuario=True))

        db.commit()


app = Flask(__name__)

def now_lima():
    return datetime.now(APP_TZ).replace(tzinfo=None)

def start_for_scale(scale):
    now = now_lima()
    options = {
        "1h": now - timedelta(hours=1),
        "6h": now - timedelta(hours=6),
        "12h": now - timedelta(hours=12),
        "24h": now - timedelta(hours=24),
        "3d": now - timedelta(days=3),
        "7d": now - timedelta(days=7),
    }
    return options.get(scale, now - timedelta(hours=24))

def fmt_dt(dt):
    return dt.strftime("%Y-%m-%d %H:%M:%S")

def parse_client_datetime(value):
    """
    Convierte la fecha enviada por el ESP32 a hora de Perú.
    Si el ESP32 no envía una fecha válida, usa la hora de recepción.
    """
    if not value:
        return now_lima()

    try:
        text = str(value).strip()

        # Permite formato ISO con Z.
        if text.endswith("Z"):
            text = text[:-1] + "+00:00"

        dt = datetime.fromisoformat(text)

        # Si viene con zona horaria, convertir a America/Lima y
        # guardar como datetime sin tzinfo, igual que el resto de la BD.
        if dt.tzinfo is not None:
            dt = dt.astimezone(APP_TZ).replace(tzinfo=None)

        return dt

    except (ValueError, TypeError):
        return now_lima()

def get_reset_time(pet_id):
    with SessionLocal() as db:
        r = db.execute(
            select(ReinicioSesion)
            .where(ReinicioSesion.mascota_id == pet_id)
            .limit(1)
        ).scalar_one_or_none()
    return r.fecha_hora if r else None

def visible_start(pet_id, scale=None, today_only=False):
    starts = []
    if scale:
        starts.append(start_for_scale(scale))
    if today_only:
        starts.append(now_lima().replace(hour=0, minute=0, second=0, microsecond=0))
    reset_time = get_reset_time(pet_id)
    if reset_time:
        starts.append(reset_time)
    return max(starts) if starts else reset_time

def valid_api_key():
    return request.headers.get("X-API-Key", "") == API_KEY

HTML = r"""<!doctype html>
<html lang="es">
<head>
<!-- V2.2.2 CORREGIDA -->
<!-- V2.2 OPCIONES AVANZADAS -->
<!-- Comedero IoT V2.1.1 limpio -->
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1,maximum-scale=5,user-scalable=yes">
<title>Monitoreo de alimentación</title>

<style>
:root{
  --bg:#f3f5f7;
  --card:#ffffff;
  --text:#1f2937;
  --muted:#6b7280;
  --border:#e5e7eb;
  --accent:#2563eb;
}
*{box-sizing:border-box}
html,body{
  margin:0;padding:0;width:100%;max-width:100%;overflow-x:hidden;
}
body{
  font-family:system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",Roboto,Arial,sans-serif;
  background:var(--bg);color:var(--text);
  -webkit-text-size-adjust:100%;text-size-adjust:100%;
}
header{
  width:100%;background:#fff;border-bottom:1px solid var(--border);
  padding:16px;display:flex;justify-content:space-between;gap:14px;
  align-items:center;flex-wrap:wrap;
}
h1{font-size:22px;line-height:1.2;margin:0}
main{width:100%;max-width:1200px;margin:0 auto;padding:16px}
.toolbar{display:flex;gap:8px;align-items:center;flex-wrap:wrap;min-width:0}
select,button{font:inherit;font-size:16px;max-width:100%}
select{
  min-height:44px;padding:10px 12px;border:1px solid var(--border);
  border-radius:10px;background:#fff;
}
.cards{
  width:100%;display:grid;grid-template-columns:repeat(4,minmax(0,1fr));
  gap:12px;margin:16px 0;
}
.card,.panel{
  min-width:0;width:100%;background:#fff;border:1px solid var(--border);
  border-radius:16px;padding:16px;
}
.label{font-size:13px;color:var(--muted);margin-bottom:8px}
.value{font-size:30px;line-height:1.15;font-weight:750;word-break:break-word}
.small{font-size:14px}
.dot{
  display:inline-block;width:10px;height:10px;border-radius:50%;
  background:#9ca3af;margin-right:6px;
}
.panel-head{
  display:flex;justify-content:space-between;gap:12px;align-items:center;
  flex-wrap:wrap;margin-bottom:12px;
}
.panel-head h2{font-size:18px;line-height:1.3;margin:0}
.ranges{display:flex;gap:6px;flex-wrap:wrap}
.range,.action{
  border:1px solid var(--border);background:#fff;border-radius:9px;
  padding:9px 12px;cursor:pointer;min-height:44px;touch-action:manipulation;
}
.range.active{background:#1f2937;color:#fff;border-color:#1f2937}
.action.primary{background:var(--accent);color:#fff;border-color:var(--accent)}
.action.danger{background:#fff;color:#b42318;border-color:#f1b4ae}
.action.danger:hover{background:#fff4f2}
.action.delete{background:#b42318;color:#fff;border-color:#b42318}
.action.delete:hover{background:#8f1d14}

.chart-toolbar{
  display:flex;justify-content:space-between;align-items:center;
  gap:8px;flex-wrap:wrap;margin:2px 0 8px;
}
.chart-help{font-size:12px;color:var(--muted)}
.zoom-actions{display:flex;gap:6px;flex-wrap:wrap}
.zoom-actions button{font-size:13px;min-height:38px;padding:7px 10px}

.chart-wrap{
  position:relative;width:100%;min-width:0;height:430px;
  border-top:1px solid var(--border);padding-top:12px;
  overflow:hidden;border-radius:8px;
}
canvas{
  display:block;width:100%!important;max-width:100%;height:100%!important;
  touch-action:none;
}
.tip{
  position:absolute;display:none;pointer-events:none;background:#111827;color:#fff;
  padding:8px 10px;border-radius:8px;font-size:12px;white-space:nowrap;
  z-index:10;box-shadow:0 4px 16px rgba(0,0,0,.18);
}
.actions{
  display:flex;justify-content:flex-end;gap:8px;flex-wrap:wrap;margin-top:14px;
}
.empty{color:var(--muted);font-size:14px}
.cloud{font-size:12px;color:var(--muted);margin-top:4px}
.hidden-view{display:none !important;}
.select-screen{
  min-height:70vh;
  display:flex;
  flex-direction:column;
  justify-content:center;
  align-items:center;
  gap:20px;
}
.select-card{
  background:#fff;
  border:1px solid var(--border);
  border-radius:18px;
  padding:30px;
  width:min(420px,95%);
  text-align:center;
}
.pet-choice{
  width:100%;
  margin:8px 0;
  padding:14px;
  border-radius:12px;
  border:1px solid var(--border);
  background:#fff;
  cursor:pointer;
  font-size:18px;
}
.back-btn{
  background:#fff;
}

@media(max-width:850px){
  .cards{grid-template-columns:repeat(2,minmax(0,1fr))}
  .chart-wrap{height:390px}
}
@media(max-width:600px){
  header{padding:12px;align-items:stretch}
  header>div{min-width:0}
  h1{font-size:20px}
  .toolbar{width:100%;display:grid;grid-template-columns:1fr}
  .toolbar label{font-size:13px;color:var(--muted)}
  .toolbar select,.toolbar button{width:100%}
  main{padding:10px}
  .cards{grid-template-columns:1fr;gap:10px;margin:10px 0}
  .card,.panel{padding:14px;border-radius:13px}
  .value{font-size:28px}
  .panel-head{align-items:stretch}
  .ranges{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));width:100%}
  .range{width:100%;padding:8px 5px}
  .chart-toolbar{align-items:stretch}
  .chart-help{width:100%;line-height:1.35}
  .zoom-actions{width:100%;display:grid;grid-template-columns:1fr}
  .zoom-actions button{width:100%}
  .chart-wrap{height:360px}
  .actions{display:grid;grid-template-columns:1fr}
  .actions .action{width:100%}
}
@media(max-width:360px){
  h1{font-size:18px}
  .ranges{grid-template-columns:repeat(2,minmax(0,1fr))}
  .value{font-size:25px}
  .chart-wrap{height:330px}
}

.modalOverlay{
display:none;
position:fixed;
inset:0;
background:rgba(15,23,42,.45);
backdrop-filter:blur(3px);
align-items:center;
justify-content:center;
z-index:999;
}

.modalCard{
background:white;
width:min(430px,90%);
padding:28px;
border-radius:22px;
box-shadow:0 20px 50px rgba(0,0,0,.25);
position:relative;
}

.modalCard h2{
margin-top:0;
color:#102a43;
}

.modalClose{
position:absolute;
right:18px;
top:14px;
border:none;
background:none;
font-size:28px;
cursor:pointer;
}

.modalInput{
width:100%;
padding:10px;
border:1px solid #cbd5e1;
border-radius:8px;
margin:8px 0;
box-sizing:border-box;
}

.modalButton{
margin-top:15px;
width:100%;
}

.action.danger{
background:#dc2626;
color:white;
}

.advancedCard .action{
display:block;
width:100%;
margin:10px 0;
}


.masterCard{background:#fff;border:1px solid #ddd;border-radius:14px;padding:14px;margin:12px 0;box-shadow:0 2px 8px #0001;}
.masterCard h4{margin:0 0 8px;font-size:18px;}
.masterActions{display:flex;gap:8px;flex-wrap:wrap;margin-top:12px;}
</style>
</head>
<body>
<header>
<div>
<h1>🐱 Monitoreo de alimentación</h1>
<div class="cloud">Plataforma en Internet · hora de Perú</div>
</div>
<div class="toolbar">
<label for="pet">Mascota</label>
<button id="masterMode" class="action">MODO MAESTRO</button>



</div>
</header>

<main>

<section id="selectScreen" class="select-screen">
  <div class="select-card">
    <h1>🐾 Comedero IoT</h1>
    <p>Seleccione mascota</p>
    <div id="petList"></div>
  </div>
</section>

<section id="dashboard" class="hidden-view">
<div class="actions" style="justify-content:flex-start;margin-top:0">
<button id="backPets" class="action back-btn">← REGRESAR</button>
</div>

<div class="cards">
<div class="card"><div class="label">Peso actual</div><div id="weight" class="value">0000 g</div><div class="small" style="margin-top:8px"><span id="dot" class="dot"></span><span id="status">Sin datos</span></div></div>
<div class="card"><div class="label">Consumo estimado hoy</div><div id="today" class="value">0.0 g</div></div>
<div class="card"><div class="label">Último consumo detectado</div><div id="lastc" class="value">0.0 g</div></div>
<div class="card"><div class="label">Última actualización</div><div id="lastt" class="value small">--</div></div>
</div>

<section class="panel">
<div class="panel-head">
<h2>Evolución de la masa de alimento disponible</h2>
<div class="ranges">
<button class="range" data-range="1h">1 h</button>
<button class="range" data-range="6h">6 h</button>
<button class="range" data-range="12h">12 h</button>
<button class="range active" data-range="24h">24 h</button>
<button class="range" data-range="3d">3 días</button>
<button class="range" data-range="7d">7 días</button>
</div>
</div>

<div class="chart-toolbar">
<div class="chart-help">PC: rueda para zoom y arrastra para desplazarte · Celular: pellizca con dos dedos y arrastra.</div>
<div class="zoom-actions">
<button id="resetZoom" class="action">Restablecer zoom</button>
</div>
</div>

<div class="chart-wrap" id="chartWrap">
<canvas id="chart"></canvas>
<div id="tip" class="tip"></div>
</div>
<div id="empty" class="empty" hidden>No hay mediciones en esta escala de tiempo.</div>



<div class="actions" style="justify-content:center;margin-top:18px">
<button id="advancedBtn" class="action">⚙ OPCIONES AVANZADAS</button>
</div>
</section>
</main>

<script>
(()=>{
const pet=document.createElement('select');
const petList=document.getElementById('petList');
const selectScreen=document.getElementById('selectScreen');
const dashboard=document.getElementById('dashboard');
const backPets=document.getElementById('backPets');
const weight=document.getElementById('weight');
const today=document.getElementById('today');
const lastc=document.getElementById('lastc');
const lastt=document.getElementById('lastt');
const status=document.getElementById('status');
const dot=document.getElementById('dot');
const canvas=document.getElementById('chart');
const chartWrap=document.getElementById('chartWrap');
const ctx=canvas.getContext('2d');
const tip=document.getElementById('tip');
const empty=document.getElementById('empty');
const resetZoom=document.getElementById('resetZoom');

let selected='',scale='24h',points=[];
let sessionMax=10;
let view=null;
let dragging=false,dragStart=null;
let pinchStart=null;

async function getj(url){
  const r=await fetch(url,{cache:'no-store'});
  if(!r.ok) throw new Error('HTTP '+r.status);
  return r.json();
}
function parseLocal(s){return s?new Date(s.replace(' ','T')):null;}
function fmt(v){return Math.round(Number(v)||0).toString().padStart(4,'0')+' g';}

function niceCeil(v){
  v=Math.max(1,Number(v)||1);
  const exp=Math.pow(10,Math.floor(Math.log10(v)));
  const f=v/exp;
  let n;
  if(f<=1)n=1;
  else if(f<=2)n=2;
  else if(f<=5)n=5;
  else n=10;
  return n*exp;
}

function baseView(){
  if(!points.length)return null;
  const ts=points.map(q=>q.t.getTime());
  let x0=Math.min(...ts),x1=Math.max(...ts);
  if(x0===x1){x0-=30000;x1+=30000;}
  const maxMeasured=Math.max(sessionMax,...points.map(q=>q.y),1);
  const margin=Math.max(5,maxMeasured*0.10);
  const y1=niceCeil(maxMeasured+margin);
  return {x0,x1,y0:0,y1};
}

function resetView(){
  view=baseView();
  hideTip();
  draw(points);
}

async function loadPets(){
  const arr=await getj('/api/mascotas');
  petList.innerHTML='';

  if(!arr.length){
    pet.innerHTML='<option value="">Esperando datos...</option>';
    selected='';points=[];sessionMax=10;view=null;draw([]);return;
  }
  arr.forEach(x=>{
    const b=document.createElement('button');
    b.className='pet-choice';
    b.textContent='🐱 '+x.nombre;
    b.onclick=async()=>{
      let necesitaClave=true;

      try{
        const cfg=await fetch("/api/master/config").then(r=>r.json());
        necesitaClave=!!cfg.solicitar_clave;
      }catch(e){
        necesitaClave=true;
      }

      let clave="";

      if(necesitaClave){
        clave=prompt("Ingrese clave de 4 dígitos:");
        if(clave===null)return;

        if(!/^\\d{4}$/.test(clave.trim())){
          alert("Ingrese una clave válida de 4 dígitos");
          return;
        }
      }

      const r=await fetch("/api/mascota/verificar",{
        method:"POST",
        headers:{"Content-Type":"application/json"},
        body:JSON.stringify({mascota_id:x.id,clave})
      });

      const j=await r.json();

      if(!j.ok){
        alert("Clave incorrecta");
        return;
      }

      selected=x.id;
      selectScreen.classList.add('hidden-view');
      dashboard.classList.remove('hidden-view');
      updateAll(true);
    };
    petList.appendChild(b);
  });
  if(!selected){
    dashboard.classList.add('hidden-view');
  }
}

async function updateCurrent(){
  if(!selected)return;
  const d=await getj('/api/ultimo/'+encodeURIComponent(selected));
  weight.textContent=fmt(d.peso);
  lastt.textContent=d.fecha_hora||'--';
  const t=parseLocal(d.fecha_hora);
  if(!t){status.textContent='Sin datos';dot.style.background='#9ca3af';return;}
  const sec=(Date.now()-t.getTime())/1000;
  if(sec<90){status.textContent='En línea';dot.style.background='#16a34a';}
  else if(sec<300){status.textContent='Datos recientes';dot.style.background='#d97706';}
  else{status.textContent='Sin conexión';dot.style.background='#dc2626';}
}

async function updateStats(){
  if(!selected)return;
  const d=await getj('/api/estadisticas/'+encodeURIComponent(selected));
  today.textContent=Number(d.consumo_hoy||0).toFixed(1)+' g';
  lastc.textContent=Number(d.ultimo_consumo||0).toFixed(1)+' g';
}

async function updateChart(forceReset=false){
  if(!selected){points=[];view=null;draw([]);return;}
  const d=await getj('/api/historico/'+encodeURIComponent(selected)+'?escala='+scale);
  points=d.fechas.map((f,i)=>({t:parseLocal(f),raw:f,y:Number(d.pesos[i])})).filter(q=>q.t && Number.isFinite(q.y));
  sessionMax=Number(d.maximo_sesion||0);
  if(forceReset || !view){
    view=baseView();
  }else{
    // El rango automático puede crecer si aparece un nuevo máximo,
    // pero nunca encoge mientras el usuario está observando la misma sesión.
    const b=baseView();
    if(b && view){
      view.x1=Math.max(view.x1,b.x1);
      if(view.y1 < b.y1 && Math.abs(view.y0) < 1e-9) view.y1=b.y1;
    }
  }
  draw(points);
}
async function updateAll(forceReset=false){
  await Promise.all([updateCurrent(),updateStats(),updateChart(forceReset)]);
}

function resize(){
  const dpr=Math.min(window.devicePixelRatio||1,2);
  const r=canvas.getBoundingClientRect();
  const cssW=Math.max(280,Math.floor(r.width||canvas.parentElement.clientWidth||320));
  const cssH=Math.max(240,Math.floor(r.height||360));
  canvas.width=Math.round(cssW*dpr);
  canvas.height=Math.round(cssH*dpr);
  ctx.setTransform(dpr,0,0,dpr,0,0);
}

function tickStep(maxVal){
  const target=Math.max(1,maxVal/6);
  return niceCeil(target);
}

function draw(data){
  resize();
  const W=canvas.clientWidth,H=canvas.clientHeight;
  ctx.clearRect(0,0,W,H);
  const p={l:58,r:28,t:22,b:42},pw=Math.max(1,W-p.l-p.r),ph=Math.max(1,H-p.t-p.b);
  ctx.font='12px system-ui';ctx.lineWidth=1;

  if(!data.length || !view){
    empty.hidden=false;
    ctx.fillStyle='#6b7280';ctx.textAlign='center';ctx.textBaseline='top';
    ctx.fillText('Tiempo',p.l+pw/2,H-20);
    canvas._geom=null;
    return;
  }
  empty.hidden=true;

  const xSpan=Math.max(1,view.x1-view.x0);
  const ySpan=Math.max(0.01,view.y1-view.y0);
  const X=t=>p.l+((t.getTime()-view.x0)/xSpan)*pw;
  const Y=v=>p.t+ph-((v-view.y0)/ySpan)*ph;

  const step=tickStep(ySpan);
  const first=Math.ceil(view.y0/step)*step;
  ctx.textBaseline='middle';
  for(let g=first;g<=view.y1+step*0.001;g+=step){
    const y=Y(g);
    if(y<p.t-1||y>p.t+ph+1)continue;
    ctx.strokeStyle='#e5e7eb';ctx.beginPath();ctx.moveTo(p.l,y);ctx.lineTo(W-p.r,y);ctx.stroke();
    ctx.fillStyle='#6b7280';ctx.textAlign='right';
    const label=(Math.abs(g)>=100 || Math.abs(g-Math.round(g))<0.01)?Math.round(g).toString():g.toFixed(1);
    ctx.fillText(label+' g',p.l-8,y);
  }

  ctx.fillStyle='#6b7280';ctx.textAlign='center';ctx.textBaseline='top';
  ctx.fillText('Tiempo',p.l+pw/2,H-20);

  // Línea escalonada: mantiene el último peso hasta que aparece un nuevo nivel.
  ctx.save();
  ctx.beginPath();
  ctx.rect(p.l,p.t,pw,ph);
  ctx.clip();

  ctx.strokeStyle='#2563eb';ctx.lineWidth=2;ctx.beginPath();
  let started=false;
  let prev=null;
  data.forEach(q=>{
    const x=X(q.t),y=Y(q.y);
    if(!started){
      ctx.moveTo(x,y);started=true;
    }else{
      const py=Y(prev.y);
      ctx.lineTo(x,py); // tramo horizontal
      ctx.lineTo(x,y);  // cambio vertical
    }
    prev=q;
  });
  ctx.stroke();

  if(data.length<=500){
    ctx.fillStyle='#2563eb';
    data.forEach(q=>{
      const x=X(q.t),y=Y(q.y);
      if(x<p.l-4||x>W-p.r+4||y<p.t-4||y>p.t+ph+4)return;
      ctx.beginPath();ctx.arc(x,y,2.6,0,Math.PI*2);ctx.fill();
    });
  }
  ctx.restore();

  canvas._geom={X,Y,p,pw,ph,W,H};
}

function hideTip(){tip.style.display='none';}

function placeTip(best){
  if(!best)return hideTip();
  const hora=best.q.t.toLocaleTimeString([],{hour:'2-digit',minute:'2-digit',second:'2-digit'});
  tip.textContent=hora+' · '+best.q.y.toFixed(1)+' g';
  tip.style.display='block';

  // Primero medimos el tooltip, luego lo limitamos dentro del recuadro.
  const tw=tip.offsetWidth,th=tip.offsetHeight;
  const cw=chartWrap.clientWidth,ch=chartWrap.clientHeight;
  let left=best.x-tw/2;
  let top=best.y-th-12;

  left=Math.max(6,Math.min(cw-tw-6,left));
  if(top<6) top=Math.min(ch-th-6,best.y+12);
  top=Math.max(6,Math.min(ch-th-6,top));

  tip.style.left=left+'px';
  tip.style.top=top+'px';
}

function nearestPoint(clientX,clientY=null,maxDist=36){
  if(!points.length||!canvas._geom)return null;
  const r=canvas.getBoundingClientRect();
  const mx=clientX-r.left,my=clientY==null?null:clientY-r.top,G=canvas._geom;
  let best=null,bd=Infinity;
  points.forEach(q=>{
    const x=G.X(q.t),y=G.Y(q.y);
    const d=my==null?Math.abs(x-mx):Math.hypot(x-mx,(y-my)*0.35);
    if(d<bd){bd=d;best={q,x,y};}
  });
  return best && bd<=maxDist ? best : null;
}

canvas.addEventListener('mousemove',ev=>{
  if(dragging)return;
  const best=nearestPoint(ev.clientX,ev.clientY,32);
  if(best)placeTip(best); else hideTip();
});
canvas.addEventListener('mouseleave',()=>{if(!dragging)hideTip();});

function clampView(){
  if(!view)return;
  const bx=baseView();
  if(!bx)return;
  const minXSpan=Math.max(1000,(bx.x1-bx.x0)/200);
  const minYSpan=Math.max(0.2,bx.y1/500);
  if(view.x1-view.x0<minXSpan)view.x1=view.x0+minXSpan;
  if(view.y1-view.y0<minYSpan)view.y1=view.y0+minYSpan;
  // Permitimos desplazamiento, pero evitamos perder completamente la zona de datos.
  const padX=(bx.x1-bx.x0)*0.5;
  if(view.x1<bx.x0-padX){const d=(bx.x0-padX)-view.x1;view.x0+=d;view.x1+=d;}
  if(view.x0>bx.x1+padX){const d=view.x0-(bx.x1+padX);view.x0-=d;view.x1-=d;}
  // El peso físico no puede ser negativo. El eje vertical se limita
  // estrictamente a Y >= 0 incluso al hacer zoom o arrastrar.
  if(view.y0 < 0){
    const spanY = Math.max(minYSpan, view.y1 - view.y0);
    view.y0 = 0;
    view.y1 = spanY;
  }
  if(view.y1 <= view.y0){
    view.y1 = view.y0 + minYSpan;
  }
}

function zoomAt(factor,cx,cy){
  if(!view||!canvas._geom)return;
  const r=canvas.getBoundingClientRect(),G=canvas._geom;
  const px=Math.max(G.p.l,Math.min(r.width-G.p.r,cx-r.left));
  const py=Math.max(G.p.t,Math.min(r.height-G.p.b,cy-r.top));
  const fx=(px-G.p.l)/G.pw;
  const fy=1-(py-G.p.t)/G.ph;
  const ax=view.x0+fx*(view.x1-view.x0);
  const ay=view.y0+fy*(view.y1-view.y0);
  view.x0=ax-(ax-view.x0)*factor;
  view.x1=ax+(view.x1-ax)*factor;
  view.y0=ay-(ay-view.y0)*factor;
  view.y1=ay+(view.y1-ay)*factor;
  clampView();draw(points);hideTip();
}

canvas.addEventListener('wheel',ev=>{
  if(!view)return;
  ev.preventDefault();
  const factor=ev.deltaY<0?0.82:1.22;
  zoomAt(factor,ev.clientX,ev.clientY);
},{passive:false});

canvas.addEventListener('pointerdown',ev=>{
  if(ev.pointerType==='touch')return; // touch se maneja abajo para permitir pinch
  if(!view)return;
  dragging=true;
  dragStart={x:ev.clientX,y:ev.clientY,v:{...view}};
  canvas.setPointerCapture?.(ev.pointerId);
  hideTip();
});
canvas.addEventListener('pointermove',ev=>{
  if(!dragging||!dragStart||!canvas._geom)return;
  const G=canvas._geom;
  const dx=ev.clientX-dragStart.x,dy=ev.clientY-dragStart.y;
  const sx=(dragStart.v.x1-dragStart.v.x0)/G.pw;
  const sy=(dragStart.v.y1-dragStart.v.y0)/G.ph;
  view.x0=dragStart.v.x0-dx*sx;view.x1=dragStart.v.x1-dx*sx;
  view.y0=dragStart.v.y0+dy*sy;view.y1=dragStart.v.y1+dy*sy;
  clampView();draw(points);
});
canvas.addEventListener('pointerup',()=>{dragging=false;dragStart=null;});
canvas.addEventListener('pointercancel',()=>{dragging=false;dragStart=null;});

function touchDistance(a,b){return Math.hypot(a.clientX-b.clientX,a.clientY-b.clientY);}
function touchCenter(a,b){return {x:(a.clientX+b.clientX)/2,y:(a.clientY+b.clientY)/2};}

canvas.addEventListener('touchstart',ev=>{
  if(!view)return;
  if(ev.touches.length===2){
    ev.preventDefault();
    const c=touchCenter(ev.touches[0],ev.touches[1]);
    pinchStart={dist:touchDistance(ev.touches[0],ev.touches[1]),center:c,v:{...view}};
    dragging=false;dragStart=null;hideTip();
  }else if(ev.touches.length===1){
    const t=ev.touches[0];
    dragStart={x:t.clientX,y:t.clientY,v:{...view},touch:true,moved:false};
  }
},{passive:false});

canvas.addEventListener('touchmove',ev=>{
  if(!view||!canvas._geom)return;
  if(ev.touches.length===2 && pinchStart){
    ev.preventDefault();
    const dist=touchDistance(ev.touches[0],ev.touches[1]);
    const c=touchCenter(ev.touches[0],ev.touches[1]);
    view={...pinchStart.v};
    const factor=Math.max(0.2,Math.min(5,pinchStart.dist/Math.max(1,dist)));
    zoomAt(factor,pinchStart.center.x,pinchStart.center.y);
    // pequeño desplazamiento del centro de la pinza
    const G=canvas._geom;
    const dx=c.x-pinchStart.center.x,dy=c.y-pinchStart.center.y;
    const sx=(view.x1-view.x0)/G.pw,sy=(view.y1-view.y0)/G.ph;
    view.x0-=dx*sx;view.x1-=dx*sx;view.y0+=dy*sy;view.y1+=dy*sy;
    clampView();draw(points);
  }else if(ev.touches.length===1 && dragStart){
    ev.preventDefault();
    const t=ev.touches[0],G=canvas._geom;
    const dx=t.clientX-dragStart.x,dy=t.clientY-dragStart.y;
    if(Math.hypot(dx,dy)>6)dragStart.moved=true;
    const sx=(dragStart.v.x1-dragStart.v.x0)/G.pw;
    const sy=(dragStart.v.y1-dragStart.v.y0)/G.ph;
    view.x0=dragStart.v.x0-dx*sx;view.x1=dragStart.v.x1-dx*sx;
    view.y0=dragStart.v.y0+dy*sy;view.y1=dragStart.v.y1+dy*sy;
    clampView();draw(points);hideTip();
  }
},{passive:false});

canvas.addEventListener('touchend',ev=>{
  if(ev.touches.length<2)pinchStart=null;
  if(ev.touches.length===0 && dragStart){
    if(!dragStart.moved){
      const changed=ev.changedTouches?.[0];
      if(changed){
        const best=nearestPoint(changed.clientX,changed.clientY,42);
        if(best)placeTip(best); else hideTip();
      }
    }
    dragStart=null;
  }
},{passive:false});

canvas.addEventListener('click',ev=>{
  if(ev.pointerType==='touch')return;
  const best=nearestPoint(ev.clientX,ev.clientY,42);
  if(best)placeTip(best); else hideTip();
});

if(resetZoom) resetZoom.addEventListener('click',resetView);

const masterMode=document.getElementById('masterMode');
if(masterMode){
  masterMode.addEventListener('click',()=>{
    const modal=document.getElementById('masterModal');
    if(modal) modal.style.display='flex';
  });
}

backPets.addEventListener('click',()=>{
  dashboard.classList.add('hidden-view');
  selectScreen.classList.remove('hidden-view');
});


document.querySelectorAll('.range').forEach(b=>b.addEventListener('click',()=>{
  document.querySelectorAll('.range').forEach(x=>x.classList.remove('active'));
  b.classList.add('active');
  scale=b.dataset.range;
  view=null;
  updateChart(true);
}));
const resetData=document.getElementById('resetData');
const deleteAll=document.getElementById('deleteAll');

const refreshBtn=document.getElementById('refresh');
if(refreshBtn) refreshBtn.addEventListener('click',()=>updateAll(false));

if(resetData) resetData.addEventListener('click',async()=>{
  if(!selected)return;
  const nombre=selected||'esta mascota';
  const ok=confirm(
    '¿Reiniciar la toma de datos de '+nombre+'?'+
    'La gráfica y los cálculos visibles comenzarán desde cero. '+
    'Los registros históricos NO se eliminarán de la base de datos.'
  );
  if(!ok)return;

  resetData.disabled=true;
  resetData.textContent='Reiniciando...';
  try{
    const r=await fetch('/api/reiniciar/'+encodeURIComponent(selected),{method:'POST'});
    if(!r.ok)throw new Error('HTTP '+r.status);
    today.textContent='0.0 g';
    lastc.textContent='0.0 g';
    points=[];sessionMax=10;view=null;draw([]);
    await updateAll(true);
    alert('Toma de datos reiniciada correctamente.');
  }catch(e){
    alert('No se pudo reiniciar la toma de datos.');
  }finally{
    resetData.disabled=false;
    resetData.textContent='Reiniciar toma de datos';
  }
});

if(deleteAll) deleteAll.addEventListener('click',async()=>{
  if(!selected)return;
  const nombre=selected||'esta mascota';
  const ok=confirm(
    '¿Estás seguro de eliminar todo el registro?'+
    'Se eliminará permanentemente TODO el historial de '+nombre+'. '+
    'Esta acción no se puede deshacer.'
  );
  if(!ok)return;

  deleteAll.disabled=true;
  deleteAll.textContent='Eliminando...';
  try{
    const r=await fetch('/api/eliminar_historial/'+encodeURIComponent(selected),{method:'DELETE'});
    if(!r.ok)throw new Error('HTTP '+r.status);
    weight.textContent='-- g';today.textContent='0.0 g';lastc.textContent='0.0 g';
    lastt.textContent='--';status.textContent='Sin datos';dot.style.background='#9ca3af';
    points=[];sessionMax=10;view=null;draw([]);
    await loadPets();
    alert('Todo el historial fue eliminado correctamente.');
  }catch(e){
    alert('No se pudo eliminar el historial.');
  }finally{
    deleteAll.disabled=false;
    deleteAll.textContent='Eliminar historial completo';
  }
});

document.addEventListener('DOMContentLoaded',()=>{

const advancedBtn=document.getElementById('advancedBtn');
const advancedPanel=document.getElementById('advancedPanel');
const closeAdvanced=document.getElementById('closeAdvanced');
const verifyAdvanced=document.getElementById('verifyAdvanced');
const advancedKey=document.getElementById('advancedKey');
const excelDownload=document.getElementById('excelDownload');
const openExcel=document.getElementById('openExcel');
const excelPanel=document.getElementById('excelPanel');
const closeExcel=document.getElementById('closeExcel');
const fullSwitch=document.getElementById('fullSwitch');
const rangeSwitch=document.getElementById('rangeSwitch');
const dateBox=document.getElementById('dateBox');
const restartData=document.getElementById('restartData');
const deleteHistory=document.getElementById('deleteHistory');
const changeKey=document.getElementById('changeKey');

if(advancedBtn && advancedPanel){
 advancedBtn.onclick=()=>{
   advancedPanel.style.display='flex';
   if(advancedKey) advancedKey.value='';
 };
}

if(closeAdvanced && advancedPanel){
 closeAdvanced.onclick=()=>{
   advancedPanel.style.display='none';
   const l=document.getElementById('advancedLogin');
   const o=document.getElementById('advancedOptions');
   if(l)l.style.display='block';
   if(o)o.style.display='none';
   if(advancedKey)advancedKey.value='';
 };
}

if(openExcel && excelPanel){
 openExcel.onclick=()=>{
   advancedPanel.style.display='none';
   excelPanel.style.display='flex';
 };
}

if(closeExcel && excelPanel){
 closeExcel.onclick=()=>excelPanel.style.display='none';
}

// Acciones del panel avanzado
if(restartData){
 restartData.onclick=async()=>{
   if(!selected){
     alert('Seleccione una mascota primero.');
     return;
   }
   if(!confirm('¿Reiniciar la toma de datos de esta mascota?')) return;
   try{
     const r=await fetch('/api/reiniciar/'+encodeURIComponent(selected),{method:'POST'});
     if(!r.ok) throw new Error();
     alert('Toma de datos reiniciada correctamente.');
     await updateAll(true);
   }catch(e){
     alert('No se pudo reiniciar la toma de datos.');
   }
 };
}

if(deleteHistory){
 deleteHistory.onclick=async()=>{
   if(!selected){
     alert('Seleccione una mascota primero.');
     return;
   }
   if(!confirm('¿Eliminar todo el historial? Esta acción no se puede deshacer.')) return;
   try{
     const r=await fetch('/api/eliminar_historial/'+encodeURIComponent(selected),{method:'DELETE'});
     if(!r.ok) throw new Error();
     alert('Historial eliminado correctamente.');
     await updateAll(true);
   }catch(e){
     alert('No se pudo eliminar el historial.');
   }
 };
}

if(changeKey){
 changeKey.onclick=()=>{
   const nueva=prompt('Ingrese la nueva clave de 4 dígitos:');
   if(nueva && /^\\d{4}$/.test(nueva)){
     alert('Clave actualizada correctamente.');
   }else if(nueva!==null){
     alert('La clave debe tener 4 dígitos.');
   }
 };
}


if(verifyAdvanced){
 verifyAdvanced.onclick=async()=>{
   const key = advancedKey.value.trim();

   if(!/^\d{4}$/.test(key)){
     alert('Ingrese una clave válida de 4 dígitos');
     return;
   }

   if(!selected){
     alert('Seleccione una mascota primero.');
     return;
   }

   try{
     const r=await fetch('/api/mascota/verificar_clave_avanzada',{
       method:'POST',
       headers:{'Content-Type':'application/json'},
       body:JSON.stringify({mascota_id:selected, clave:key})
     });

     const j=await r.json();

     if(!j.ok){
       alert('Clave incorrecta');
       return;
     }

     document.getElementById('advancedLogin').style.display='none';
     document.getElementById('advancedOptions').style.display='block';
   }catch(e){
     alert('No se pudo verificar la clave.');
   }
 };
}

function setMode(full){
 if(fullSwitch && rangeSwitch && dateBox){
   fullSwitch.checked=full;
   rangeSwitch.checked=!full;
   dateBox.style.display=full?'none':'block';
 }
}

if(fullSwitch) fullSwitch.onchange=()=>setMode(true);
if(rangeSwitch) rangeSwitch.onchange=()=>setMode(false);

if(excelDownload){
 excelDownload.onclick=()=>{
  if(!selected){
    alert('No hay mascota seleccionada');
    return;
  }

  let url='';

  if(fullSwitch.checked){
    url='/api/excel/'+encodeURIComponent(selected)+'?escala=todo';
  }else{
    const ini=document.getElementById('dateStart').value;
    const fin=document.getElementById('dateEnd').value;

    if(!ini || !fin){
      alert('Seleccione fecha inicial y fecha final');
      return;
    }

    url='/api/excel_fecha/'+encodeURIComponent(selected)
       +'?inicio='+encodeURIComponent(ini)
       +'&fin='+encodeURIComponent(fin);
  }

  window.location.href=url;
 };
}

});

let resizeTimer=null;
function redrawSoon(){
  clearTimeout(resizeTimer);
  resizeTimer=setTimeout(()=>draw(points),120);
}
window.addEventListener('resize',redrawSoon);
window.addEventListener('orientationchange',()=>setTimeout(()=>draw(points),300));
if(window.visualViewport)window.visualViewport.addEventListener('resize',redrawSoon);

setInterval(()=>{if(selected){updateCurrent();updateStats();updateChart(false);}else loadPets();},10000);
loadPets().catch(console.error);
})();



// ===== V2.3.2 MODO MAESTRO ESTABLE =====
document.addEventListener("DOMContentLoaded",()=>{
const modal=document.getElementById("masterModal");
const masterMode=document.getElementById("masterMode");
const close=document.getElementById("closeMaster");
const enter=document.getElementById("masterEnter");
const save=document.getElementById("masterSave");
const addPet=document.getElementById("addPet");

function resetMaster(){
 modal.style.display="none";
 document.getElementById("masterLogin").style.display="block";
 document.getElementById("masterPanel").style.display="none";
 document.getElementById("masterKey").value="";
}

if(masterMode) masterMode.onclick=()=>{
 modal.style.display="flex";
 resetMaster();
};

if(close) close.onclick=resetMaster;

async function cargarMaestro(){
 const res=await fetch("/api/master/mascotas");
 const pets=await res.json();
 document.getElementById("masterList").innerHTML=pets.map(p=>`
 <div class="masterCard">
  <h4>🐱 ${p.nombre}</h4>
  <div>ID: ${p.mascota_id}</div>
  <div>Estado: ${p.visible?'🟢 Visible':'🔴 Oculta'}</div>
  <div class="masterActions">
   <button class="action" onclick="toggleMascota(${p.id})">${p.visible?'Ocultar':'Mostrar'}</button>
   <button class="action" onclick="editarMascota(${p.id},'${p.nombre}')">Editar</button>
   <button class="action danger" onclick="eliminarMascota(${p.id},'${p.nombre}')">Eliminar</button>
  </div>
 </div>`).join("");
}

window.toggleMascota=async(id)=>{
 await fetch('/api/master/visible/'+id,{method:'POST'});
 cargarMaestro();
};

window.editarMascota=async(id,nombre)=>{
 const nuevo=prompt("Nuevo nombre:",nombre);
 if(!nuevo)return;
 const clave=prompt("Nueva clave de 4 dígitos (vacío conserva):");
 await fetch("/api/master/editar/"+id,{
  method:"PUT",
  headers:{"Content-Type":"application/json"},
  body:JSON.stringify({nombre:nuevo,clave:clave||""})
 });
 cargarMaestro();
};

window.eliminarMascota=async(id,nombre)=>{
 if(!confirm("Eliminar definitivamente "+nombre+"?")) return;
 await fetch("/api/master/eliminar/"+id,{method:"DELETE"});
 cargarMaestro();
};

if(addPet) addPet.onclick=async()=>{
 const mascota_id=prompt("ID ESP32:");
 const nombre=prompt("Nombre:");
 const clave=prompt("Clave de 4 dígitos:");
 if(!mascota_id||!nombre||!clave)return;

 const r=await fetch("/api/master/agregar",{
 method:"POST",
 headers:{"Content-Type":"application/json"},
 body:JSON.stringify({mascota_id,nombre,clave})
 });
 const j=await r.json();
 if(!j.ok) alert(j.mensaje||"Error");
 cargarMaestro();
 await getj('/api/mascotas').then(()=>loadPets());
};

if(enter) enter.onclick=async()=>{
 const clave=document.getElementById("masterKey").value;
 const r=await fetch("/api/master/verificar",{
 method:"POST",
 headers:{"Content-Type":"application/json"},
 body:JSON.stringify({clave})
 });
 const j=await r.json();

 if(!j.ok){
  alert("Clave incorrecta");
  return;
 }

 document.getElementById("masterLogin").style.display="none";
 document.getElementById("masterPanel").style.display="block";
 cargarMaestro();
};

const keySwitch=document.getElementById("keySwitch");
if(keySwitch){
 fetch("/api/master/config").then(r=>r.json()).then(j=>{
   keySwitch.checked=!!j.solicitar_clave;
 });
 keySwitch.onchange=async()=>{
   await fetch("/api/master/config",{
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body:JSON.stringify({solicitar_clave:keySwitch.checked})
   });
 };
}
if(save) save.onclick=resetMaster;
});
</script>

<div id="advancedPanel" class="modalOverlay">
<div class="modalCard advancedCard">

<button id="closeAdvanced" class="modalClose">×</button>

<div id="advancedLogin">
<h2>⚙ Opciones avanzadas</h2>
<p>Ingrese clave de 4 dígitos</p>
<input id="advancedKey" maxlength="4" type="password" class="modalInput">
<button id="verifyAdvanced" class="action primary modalButton">Ingresar</button>
</div>

<div id="advancedOptions" style="display:none;">
<h2>⚙ Opciones avanzadas</h2>

<button id="openExcel" class="action primary">📥 Descargar Excel</button>
<button id="restartData" class="action">🔄 Reiniciar toma de datos</button>
<button id="deleteHistory" class="action danger">🗑 Eliminar historial completo</button>
<button id="changeKey" class="action">🔑 Cambiar clave</button>

</div>
</div>
</div>

<div id="excelPanel" class="modalOverlay" style="display:none;">
<div class="modalCard">

<button id="closeExcel" class="modalClose">×</button>

<h2>📥 Descargar Excel</h2>

<label>
<input id="fullSwitch" type="radio" name="mode" checked>
Registro completo
</label>

<br>

<label>
<input id="rangeSwitch" type="radio" name="mode">
Intervalo de fechas
</label>

<div id="dateBox" style="display:none;margin-top:15px;">
<p>Fecha inicial:</p>
<input id="dateStart" type="date" class="modalInput">

<p>Fecha final:</p>
<input id="dateEnd" type="date" class="modalInput">
</div>

<button id="excelDownload" class="action primary modalButton">
Descargar
</button>

</div>
</div>


<div id="masterModal" class="modalOverlay" style="display:none">
<div class="modalCard">
<button id="closeMaster" class="modalClose">×</button>
<h2>⚙ MODO MAESTRO</h2>

<div id="masterLogin">
<p>Ingrese clave maestra</p>
<input id="masterKey" type="password" maxlength="4" class="modalInput">
<button id="masterEnter" class="action primary modalButton">Ingresar</button>
</div>

<div id="masterPanel" style="display:none">
<h3>Mascotas registradas</h3>
<div id="masterList"></div>
<button id="addPet" class="action primary">＋ AGREGAR MASCOTA</button>
<div style="margin:18px 0;padding:12px;border:1px solid #e5e7eb;border-radius:12px">
<label style="display:flex;justify-content:space-between;align-items:center">
<span>🔐 Solicitar clave de usuario</span>
<input id="keySwitch" type="checkbox" checked>
</label>
</div>
<button id="masterSave" class="action">GUARDAR Y SALIR</button>
</div>

</div>
</div>

<script>
// V2.3.6 - binding seguro del acceso MODO MAESTRO
(function(){
 function iniciarModoMaestro(){
   const btn=document.getElementById("masterMode");
   const modal=document.getElementById("masterModal");
   const close=document.getElementById("closeMaster");
   const key=document.getElementById("masterKey");
   const login=document.getElementById("masterLogin");
   const panel=document.getElementById("masterPanel");

   if(!btn || !modal) return;

   const cerrar=function(){
     modal.style.display="none";
     if(login) login.style.display="block";
     if(panel) panel.style.display="none";
     if(key) key.value="";
   };

   btn.onclick=function(e){
     e.preventDefault();
     cerrar();
     modal.style.display="flex";
   };

   if(close) close.onclick=cerrar;
 }

 if(document.readyState==="loading"){
   document.addEventListener("DOMContentLoaded", iniciarModoMaestro);
 }else{
   iniciarModoMaestro();
 }
})();
</script>

</body>
</html>"""

@app.get("/")
def home():
    return Response(HTML, mimetype="text/html")

@app.get("/health")
def health():
    return jsonify(status="ok")

@app.post("/api/peso")
def receive_weight():
    if not valid_api_key():
        return jsonify(error="API key inválida"), 401

    d = request.get_json(silent=True) or {}
    pet_id = str(d.get("mascota_id", "")).strip()
    name = str(d.get("nombre", "")).strip()

    try:
        weight = float(d.get("peso"))
    except Exception:
        return jsonify(error="Peso inválido"), 400

    if not pet_id or not name:
        return jsonify(error="Faltan mascota_id o nombre"), 400

    weight = max(0.0, min(1000.0, weight))
    timestamp = parse_client_datetime(d.get("fecha_hora"))

    with SessionLocal() as db:
        pet = db.execute(
            select(Mascota).where(Mascota.mascota_id == pet_id)
        ).scalar_one_or_none()

        if pet is None:
            db.add(Mascota(
                mascota_id=pet_id,
                nombre=name,
                clave_hash=hash_clave("1234"),
                visible=True,
                fecha_creacion=now_lima()
            ))
            db.commit()

        # Evita duplicar una medición si el ESP32 la reintenta después de
        # que el servidor ya la guardó pero la respuesta HTTP se perdió.
        existing = db.execute(
            select(Medicion)
            .where(Medicion.mascota_id == pet_id)
            .where(Medicion.fecha_hora == timestamp)
            .limit(1)
        ).scalar_one_or_none()

        if existing is None:
            db.add(Medicion(
                mascota_id=pet_id,
                nombre=name,
                fecha_hora=timestamp,
                peso=weight
            ))
            db.commit()
        else:
            weight = float(existing.peso)

    return jsonify(
        estado="ok",
        peso=weight,
        fecha_hora=fmt_dt(timestamp)
    )

@app.get("/api/mascotas")
def pets():
    with SessionLocal() as db:
        rows = db.execute(
            select(Mascota).where(
                Mascota.visible == True,
                Mascota.mascota_id != "mascota01"
            )
        ).scalars().all()

        return [
            {
                "id": r.mascota_id,
                "nombre": r.nombre
            }
            for r in rows
        ]

        return jsonify(mascotas)

@app.get("/api/ultimo/<pet_id>")
def latest(pet_id):
    with SessionLocal() as db:
        r = db.execute(
            select(Medicion)
            .where(Medicion.mascota_id == pet_id)
            .order_by(Medicion.fecha_hora.desc(), Medicion.id.desc())
            .limit(1)
        ).scalar_one_or_none()

    if not r:
        return jsonify(nombre="", fecha_hora="", peso=0)

    return jsonify(
        nombre=r.nombre,
        fecha_hora=fmt_dt(r.fecha_hora),
        peso=r.peso
    )

@app.get("/api/historico/<pet_id>")
def history(pet_id):
    scale = request.args.get("escala", "24h")
    start = visible_start(pet_id, scale=scale)
    session_start = get_reset_time(pet_id)

    with SessionLocal() as db:
        stmt = select(Medicion).where(Medicion.mascota_id == pet_id)
        if start:
            stmt = stmt.where(Medicion.fecha_hora >= start)
        rows = db.execute(stmt.order_by(Medicion.fecha_hora.asc())).scalars().all()

        max_stmt = select(Medicion).where(Medicion.mascota_id == pet_id)
        if session_start:
            max_stmt = max_stmt.where(Medicion.fecha_hora >= session_start)
        session_rows = db.execute(max_stmt).scalars().all()

    max_session = max((float(r.peso) for r in session_rows), default=0.0)

    # Conservamos cambios reales y siempre el último punto para prolongar
    # horizontalmente el último nivel en la gráfica escalonada.
    changed = []
    last_weight = None
    for r in rows:
        w = float(r.peso)
        if last_weight is None or abs(w - last_weight) >= 0.5:
            changed.append(r)
            last_weight = w

    if rows:
        if not changed or changed[-1].id != rows[-1].id:
            changed.append(rows[-1])

    if len(changed) > 1200:
        step = max(1, len(changed)//1200)
        sampled = changed[::step]
        if sampled[-1].id != changed[-1].id:
            sampled.append(changed[-1])
        changed = sampled

    return jsonify(
        fechas=[fmt_dt(r.fecha_hora) for r in changed],
        pesos=[r.peso for r in changed],
        maximo_sesion=round(max_session, 2)
    )

def calc_consumption(rows, threshold=2.0):
    if len(rows) < 2:
        return 0.0, 0.0

    total = 0.0
    last = 0.0
    prev = float(rows[0].peso)

    for r in rows[1:]:
        cur = float(r.peso)
        drop = prev - cur
        if drop >= threshold:
            total += drop
            last = drop
        prev = cur

    return round(total, 1), round(last, 1)

@app.get("/api/estadisticas/<pet_id>")
def stats(pet_id):
    start = visible_start(pet_id, today_only=True)

    with SessionLocal() as db:
        stmt = select(Medicion).where(Medicion.mascota_id == pet_id)
        if start:
            stmt = stmt.where(Medicion.fecha_hora >= start)
        rows = db.execute(stmt.order_by(Medicion.fecha_hora.asc())).scalars().all()

    total, last = calc_consumption(rows)
    return jsonify(consumo_hoy=total, ultimo_consumo=last)

@app.delete("/api/eliminar_historial/<pet_id>")
def delete_history(pet_id):
    with SessionLocal() as db:
        db.query(Medicion).filter(
            Medicion.mascota_id == pet_id
        ).delete(synchronize_session=False)

        db.query(ReinicioSesion).filter(
            ReinicioSesion.mascota_id == pet_id
        ).delete(synchronize_session=False)

        db.commit()

    return jsonify(
        estado="ok",
        mensaje="Historial eliminado completamente"
    )

@app.post("/api/reiniciar/<pet_id>")
def reset_session(pet_id):
    timestamp = now_lima()
    with SessionLocal() as db:
        r = db.execute(
            select(ReinicioSesion)
            .where(ReinicioSesion.mascota_id == pet_id)
            .limit(1)
        ).scalar_one_or_none()
        if r:
            r.fecha_hora = timestamp
        else:
            db.add(ReinicioSesion(mascota_id=pet_id, fecha_hora=timestamp))
        db.commit()

    return jsonify(estado="ok", mensaje="Toma de datos reiniciada", fecha_hora=fmt_dt(timestamp))


@app.post("/api/mascota/verificar")
def verificar_mascota():
    data=request.get_json() or {}
    pet_id=str(data.get("mascota_id", ""))
    clave=str(data.get("clave", ""))
    with SessionLocal() as db:
        m=db.query(Mascota).filter(Mascota.mascota_id==pet_id).first()
        if not m:
            return jsonify(ok=False)
        cfg=db.query(Configuracion).first()
        if cfg and not cfg.solicitar_clave_usuario:
            return jsonify(ok=True)
        return jsonify(ok=(hash_clave(clave)==m.clave_hash))


@app.post("/api/mascota/verificar_clave_avanzada")
def verificar_clave_avanzada():
    data=request.get_json() or {}
    pet_id=str(data.get("mascota_id",""))
    clave=str(data.get("clave","")).strip()
    with SessionLocal() as db:
        m=db.query(Mascota).filter(Mascota.mascota_id==pet_id).first()
        if not m:
            return jsonify(ok=False)
        return jsonify(ok=(hash_clave(clave)==m.clave_hash))

@app.post("/api/master/verificar")
def master_verificar():
    data=request.get_json() or {}
    clave=str(data.get("clave",""))
    with SessionLocal() as db:
        admin=db.query(Administrador).first()
        if not admin:
            admin=Administrador(clave_maestra_hash=hash_clave("9999"))
            db.add(admin); db.commit()
        return jsonify(ok=(hash_clave(clave)==admin.clave_maestra_hash))

@app.get("/api/master/mascotas")
def master_mascotas():
    with SessionLocal() as db:
        datos=db.query(Mascota).filter(Mascota.mascota_id != "mascota01").all()
        return jsonify([{
            "id":m.id,"mascota_id":m.mascota_id,
            "nombre":m.nombre,"visible":m.visible
        } for m in datos])

@app.post("/api/master/visible/<int:id>")
def master_visible(id):
    with SessionLocal() as db:
        m=db.get(Mascota,id)
        if not m:return jsonify(ok=False)
        m.visible=not m.visible
        db.commit()
        return jsonify(ok=True)


@app.post("/api/master/agregar")
def master_agregar():
    data=request.get_json(silent=True) or {}
    mascota_id=str(data.get("mascota_id","")).strip()
    nombre=str(data.get("nombre","")).strip()
    clave=str(data.get("clave","")).strip()

    if not mascota_id or not nombre or not(clave.isdigit() and len(clave)==4):
        return jsonify(ok=False,mensaje="Datos inválidos"),400

    with SessionLocal() as db:
        if db.execute(select(Mascota).where(Mascota.mascota_id==mascota_id)).scalar_one_or_none():
            return jsonify(ok=False,mensaje="ID ya existe"),400

        db.add(Mascota(
            mascota_id=mascota_id,
            nombre=nombre,
            clave_hash=hash_clave(clave),
            visible=True,
            fecha_creacion=now_lima()
        ))
        db.commit()
    return jsonify(ok=True)


@app.put("/api/master/editar/<int:id>")
def master_editar(id):
    data=request.get_json(silent=True) or {}
    with SessionLocal() as db:
        m=db.get(Mascota,id)
        if not m:
            return jsonify(ok=False),404

        nombre=str(data.get("nombre","")).strip()
        clave=str(data.get("clave","")).strip()

        if nombre:
            m.nombre=nombre
        if clave:
            if not(clave.isdigit() and len(clave)==4):
                return jsonify(ok=False,mensaje="Clave inválida"),400
            m.clave_hash=hash_clave(clave)

        db.commit()
    return jsonify(ok=True)


@app.delete("/api/master/eliminar/<int:id>")
def master_eliminar(id):
    with SessionLocal() as db:
        m=db.get(Mascota,id)
        if not m:
            return jsonify(ok=False),404

        db.query(Medicion).filter(Medicion.mascota_id==m.mascota_id).delete(
            synchronize_session=False
        )
        db.delete(m)
        db.commit()

    return jsonify(ok=True)


@app.get("/api/master/config")
def master_config():
    with SessionLocal() as db:
        cfg=db.query(Configuracion).first()
        return jsonify(solicitar_clave=bool(cfg.solicitar_clave_usuario) if cfg else True)



@app.post("/api/master/editar-clave/<pet_id>")
def editar_clave_usuario(pet_id):
    data=request.get_json() or {}
    clave=str(data.get("clave",""))
    if not clave.isdigit() or len(clave)!=4:
        return jsonify(ok=False,mensaje="Clave inválida"),400
    with SessionLocal() as db:
        m=db.query(Mascota).filter(Mascota.mascota_id==pet_id).first()
        if not m:
            return jsonify(ok=False),404
        m.clave_hash=hash_clave(clave)
        db.commit()
    return jsonify(ok=True)

@app.post("/api/master/config")
def guardar_master_config():
    data=request.get_json() or {}
    with SessionLocal() as db:
        cfg=db.query(Configuracion).first()
        if not cfg:
            cfg=Configuracion()
            db.add(cfg)
        cfg.solicitar_clave_usuario=bool(data.get("solicitar_clave",True))
        db.commit()
    return jsonify(ok=True)


@app.get("/api/excel/<pet_id>")
def excel(pet_id):
    scale = request.args.get("escala", "todo")

    with SessionLocal() as db:
        stmt = select(Medicion).where(Medicion.mascota_id == pet_id)
        if scale != "todo":
            start = visible_start(pet_id, scale=scale)
            if start:
                stmt = stmt.where(Medicion.fecha_hora >= start)
        rows = db.execute(stmt.order_by(Medicion.fecha_hora.asc())).scalars().all()

    if not rows:
        return Response("No hay datos para exportar.", status=404, mimetype="text/plain")

    name = rows[0].nombre
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro"

    headers = [
        "Mascota", "Fecha", "Hora", "Peso alimento (g)",
        "Cambio de peso (g)", "Consumo estimado (g)"
    ]
    ws.append(headers)

    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="D9EAF7")
        c.alignment = Alignment(horizontal="center")

    prev = None
    for r in rows:
        cur = float(r.peso)
        change = 0.0 if prev is None else cur - prev
        consumption = abs(change) if change <= -2.0 else 0.0

        ws.append([
            name,
            r.fecha_hora.strftime("%d/%m/%Y"),
            r.fecha_hora.strftime("%H:%M:%S"),
            round(cur, 1),
            round(change, 1),
            round(consumption, 1)
        ])

        prev = cur

    for i, width in enumerate([18, 14, 12, 20, 20, 24], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    f = BytesIO()
    wb.save(f)
    f.seek(0)

    stamp = now_lima().strftime("%Y%m%d_%H%M")
    suffix = "completo" if scale == "todo" else scale
    filename = f"{name}_{suffix}_{stamp}.xlsx"

    return send_file(
        f,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/excel_fecha/<pet_id>")
def excel_fecha(pet_id):
    inicio=request.args.get("inicio")
    fin=request.args.get("fin")
    try:
        ini=datetime.strptime(inicio,"%Y-%m-%d")
        fin_dt=datetime.strptime(fin,"%Y-%m-%d")+timedelta(days=1)
    except:
        return Response("Fechas inválidas",status=400)

    with SessionLocal() as db:
        rows=db.execute(
            select(Medicion)
            .where(Medicion.mascota_id==pet_id)
            .where(Medicion.fecha_hora>=ini)
            .where(Medicion.fecha_hora<fin_dt)
            .order_by(Medicion.fecha_hora.asc())
        ).scalars().all()

    if not rows:
        return Response("No hay datos",status=404)

    wb=Workbook()
    ws=wb.active
    ws.append(["Mascota","Fecha","Hora","Peso alimento (g)"])
    for r in rows:
        ws.append([r.nombre,r.fecha_hora.strftime("%d/%m/%Y"),r.fecha_hora.strftime("%H:%M:%S"),r.peso])

    f=BytesIO()
    wb.save(f)
    f.seek(0)

    return send_file(f,as_attachment=True,
        download_name=f"{pet_id}_rango.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")



inicializar_plataforma()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    app.run(host="0.0.0.0", port=port, debug=False)
